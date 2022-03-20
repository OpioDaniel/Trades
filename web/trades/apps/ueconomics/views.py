import openpyxl
from django.shortcuts import render
from .data_processing import AlgoUE
from .models import Product, Source, YearData
from openpyxl import workbook, load_workbook
import pandas as pd

def index(request):
    title = 'UBOS Export and Import'
    pdu = AlgoUE()
    df_, product_, source_ = pdu.get_data()
    return render(request, 'ueconomics/index_.html', {'title': title, 'df': df_,
                                                     'products': product_,
                                                     'sources': source_
                                                     })


def get_source_data(request):
    source_name = request.POST.get('source')
    source_ = Source.objects.get(type=source_name)
    products_ = Product.objects.all()
    # print(products_)
    p_ = products_[0]
    years = YearData.objects.filter(source=source_, product=p_)
    # print(years)

    # year_data_ = YearData.objects.filter(source__type=source_name)
    return render(request, 'ueconomics/_index_source.html', {
        'products': products_,
        'years': years,
        'source': source_
        # ,
        # 'year_data': year_data_
    })


def load_data(request):
    title = 'UBOS Export and Import'
    pdu = AlgoUE()
    print(pdu)
    pdu.upload_data_to_database()

    return render(request, 'ueconomics/update_data.html', {'title': title})

def trial(request):
    trial_ = 'DATA ON EXPORT AND IMPORT'

    path = openpyxl.load_workbook("C:\\envs\\projects\\trades\\data\\ueconomics\\datasets\\exports.xlsx")
    print(path)
    sheets = 'path.sheetnames'
    # print(path.active.title)

    sh1 = path['jean']
    index = sh1['B2'].value
    print(index)

    df_ = pd.read_excel("C:\\envs\\projects\\trades\\data\\ueconomics\\datasets\\exports.xlsx")
    print(df_)
    return render(request, 'ueconomics/trial.html',{'trial':trial_, 'index_':index,
                                                    'df':df_


                                                    })